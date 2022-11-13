import numpy as np
import cv2
import timeit
from openpyxl import load_workbook

def preprocessing(num, q):
    #exdata = load_ws.cell(num, 1).value

    if q == 0:
        img_path = 'D:/EO'+str(num)+'.png'
    else:
        img_path = 'D:/IR'+str(num)+'.png'

    img = cv2.imread(img_path)
    img = cv2.resize(img, dsize=(640, 480), interpolation=cv2.INTER_AREA)

    height, width, channels = img.shape
    blob = cv2.dnn.blobFromImage(img, 1.0 / 256, (640, 480), (0, 0, 0), swapRB=True, crop=False)

    yolo_model = cv2.dnn.readNet('D:/tiny2.weights', 'D:/tiny2.cfg')
    layer_names = yolo_model.getLayerNames()
    out_layers = [layer_names[i[0] - 1] for i in yolo_model.getUnconnectedOutLayers()]

    yolo_model.setInput(blob)
    output3 = yolo_model.forward(out_layers)

    return output3, height, width, channels, img

'''
filename = 'D:/AI/image_data.xlsx'

load_wb = load_workbook(filename, data_only=True)
load_ws = load_wb['sheet1']
'''

f=open('D:/names.txt', 'r')

classes = [line.strip() for line in f.readlines()]

colors=np.random.uniform(0,255,size=(80,3))
time0 = timeit.default_timer()

index_img = 3 # 2 ~ 3
step = [0,1,2]

for st in step:
    for num in range(1,index_img):

        if st == 2:
            img1 = cv2.imread('D:/EO_r' + str(num) + '.png')
            img2 = cv2.imread('D:/IR_r' + str(num) + '.png')

            a = 0.5
            b = 1.0 - a
            dst = cv2.addWeighted(img1, a, img2, b, 0)
            cv2.imwrite('D:/SUM_r' + str(num) + '.png', dst)

        else:
            output3, height, width, channels, img = preprocessing(num, st)

            class_ids,confidences,boxes=[],[],[]

            for output in output3:
                for vec85 in output:
                    scores=vec85[5:]
                    class_id=np.argmax(scores)
                    confidence=scores[class_id]

                    if confidence > 0.5:
                        centerx,centery=int(vec85[0]*width),int(vec85[1]*height)
                        w,h=int(vec85[2]*width),int(vec85[3]*height)
                        x,y=int(centerx-w/2),int(centery-h/2)
                        boxes.append([x,y,w,h])
                        confidences.append(float(confidence))
                        class_ids.append(class_id)
                
            indexes=cv2.dnn.NMSBoxes(boxes,confidences,0.5,0.4)

            for i in range(len(boxes)):
                if i in indexes:
                    x,y,w,h=boxes[i]
                    text=str(classes[class_ids[i]])+'%.3f'%confidences[i]
                    cv2.rectangle(img,(x,y),(x+w,y+h),colors[class_ids[i]],2)
                    cv2.putText(img,text,(x,y+30),cv2.FONT_HERSHEY_PLAIN,2,colors[class_ids[i]],2)


        if st == 0:
            img_name = 'D:/EO_r'+str(num)+'.png'
            cv2.imwrite(img_name, img)
            time1 = timeit.default_timer()

        elif st == 1 :
            img_name = 'D:/IR_r' + str(num) + '.png'
            cv2.imwrite(img_name, img)
            time2 = timeit.default_timer()


time6 = timeit.default_timer()

print("cmos %f초 " %(time1 - time0))
print("열상 %f초 " %(time2 - time1))
print("결과 %f초 " %(time6 - time0))